#!/usr/bin/env python3
"""
Israel Job Scraper Web App
===========================
Streamlit UI with two tabs:
  - AllJobs.co.il: scrape, translate, and export job listings
  - Telegram Jobs: read Hebrew job posts from a Telegram group, translate, and export
"""

import io
import os
import re
import sys
import time
from datetime import datetime

import openpyxl
import streamlit as st
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

load_dotenv()  # loads TELEGRAM_API_ID, TELEGRAM_API_HASH, TELEGRAM_PHONE, ANTHROPIC_API_KEY from .env

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from alljobs_scraper import CATEGORIES, JOB_TYPES, REGIONS, AllJobsScraper
from telegram_scraper import TelegramJobFetcher
import whatsapp_scraper as wa


# ─────────────────────────────────────────────
# Translation helpers (shared by both tabs)
# ─────────────────────────────────────────────

@st.cache_resource(show_spinner=False)
def get_translator():
    try:
        from deep_translator import GoogleTranslator
        GoogleTranslator(source="iw", target="en").translate("שלום")
        return True
    except Exception:
        return False


def translate_text(text: str) -> str:
    from deep_translator import GoogleTranslator
    if not text or not text.strip():
        return text
    try:
        chunk = text[:4800]
        result = GoogleTranslator(source="iw", target="en").translate(chunk)
        return result or text
    except Exception:
        return text


# ─────────────────────────────────────────────
# Hebrew date parser (AllJobs tab)
# ─────────────────────────────────────────────

def parse_days_ago(date_str: str) -> int:
    """Parse Hebrew relative date string → approximate days ago."""
    if not date_str:
        return 999
    s = date_str.strip()
    if any(w in s for w in ["היום", "שעה", "שעות", "דקות", "דקה", "רגע"]):
        return 0
    if "אתמול" in s:
        return 1
    m = re.search(r"(\d+)\s*ימים", s)
    if m:
        return int(m.group(1))
    if "יום" in s:
        return 1
    m = re.search(r"(\d+)\s*שבועות", s)
    if m:
        return int(m.group(1)) * 7
    if "שבוע" in s:
        return 7
    m = re.search(r"(\d+)\s*חודשים", s)
    if m:
        return int(m.group(1)) * 30
    if "חודש" in s:
        return 30
    return 999


# ─────────────────────────────────────────────
# Excel export (shared by both tabs)
# ─────────────────────────────────────────────

def jobs_to_excel(jobs: list[dict], translated: bool, source_col: str = "direct_url") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jobs"

    headers = ["Title", "Company", "Location", "Job Type", "Posted", "Description", "Requirements", "Link"]
    col_widths = [42, 32, 26, 22, 14, 65, 55, 14]

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center")

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    alt_fill = PatternFill("solid", fgColor="D6E4F0")

    for row_i, job in enumerate(jobs, 2):
        def get(field, _job=job):
            en_key = f"{field}_en"
            if translated and _job.get(en_key):
                return _job[en_key]
            return _job.get(field, "")

        ws.cell(row=row_i, column=1, value=get("title"))
        ws.cell(row=row_i, column=2, value=get("company"))
        ws.cell(row=row_i, column=3, value=get("location"))
        ws.cell(row=row_i, column=4, value=get("job_type"))
        ws.cell(row=row_i, column=5, value=job.get("date_posted", ""))
        ws.cell(row=row_i, column=6, value=get("description"))
        ws.cell(row=row_i, column=7, value=get("requirements"))

        url = job.get(source_col) or job.get("job_url") or job.get("source_url", "")
        link_cell = ws.cell(row=row_i, column=8)
        if url:
            link_cell.value = "Open →"
            link_cell.hyperlink = url
            link_cell.font = Font(color="0563C1", underline="single")
        else:
            link_cell.value = ""

        row_fill = alt_fill if row_i % 2 == 0 else None
        wrap_align = Alignment(wrap_text=True, vertical="top")
        for col in range(1, 9):
            c = ws.cell(row=row_i, column=col)
            c.alignment = wrap_align
            if row_fill and col != 8:
                c.fill = row_fill
        ws.row_dimensions[row_i].height = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────
# Shared job card renderer
# ─────────────────────────────────────────────

def render_job_cards(jobs: list[dict], translated: bool, url_field: str = "direct_url", max_desc: int = 1500):
    for job in jobs:
        def g(field, _job=job):
            en_key = f"{field}_en"
            if translated and _job.get(en_key):
                return _job[en_key]
            return _job.get(field, "") or ""

        title = g("title") or "Untitled"
        company = g("company") or "Unknown company"
        location = g("location") or ""
        date = job.get("date_posted", "")
        job_type = g("job_type") or ""
        desc = g("description") or ""
        req = g("requirements") or ""
        url = job.get(url_field) or job.get("job_url") or job.get("source_url") or ""

        label = f"**{title}** — {company}"
        if location:
            label += f" | 📍 {location}"
        if date:
            label += f" | 🕐 {date}"

        with st.expander(label):
            left, right = st.columns([4, 1])
            with left:
                if desc:
                    st.write(desc[:max_desc] + ("…" if len(desc) > max_desc else ""))
                if req:
                    st.markdown("**Requirements**")
                    st.write(req[:800] + ("…" if len(req) > 800 else ""))
            with right:
                if job_type:
                    st.markdown(f"**Type:** {job_type}")
                if date:
                    st.markdown(f"**Posted:** {date}")
                if url:
                    st.link_button("Open →", url)


# ─────────────────────────────────────────────
# Page config
# ─────────────────────────────────────────────

st.set_page_config(
    page_title="Israel Job Scraper",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
  #MainMenu {visibility: hidden;}
  footer {visibility: hidden;}
  header {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

st.title("🔍 Israel Job Scraper")
st.caption("Search AllJobs.co.il or read Hebrew job posts from Telegram / WhatsApp — translate and export as Excel.")

# ─────────────────────────────────────────────
# Sidebar
# ─────────────────────────────────────────────

with st.sidebar:
    # ── AllJobs filters ──
    st.header("🏢 AllJobs Filters")

    st.subheader("Categories (Domain)")
    cat_label_to_id = {
        f"{v['en']}  ({v['he']})": k
        for k, v in sorted(CATEGORIES.items(), key=lambda x: x[1]["en"])
    }
    selected_cat_labels = st.multiselect(
        "Select one or more domains",
        options=list(cat_label_to_id.keys()),
        placeholder="All categories...",
    )

    st.subheader("Region")
    reg_label_to_id = {
        f"{v['en']}  ({v['he']})": k
        for k, v in sorted(REGIONS.items(), key=lambda x: x[0])
    }
    selected_reg_labels = st.multiselect(
        "Select regions",
        options=list(reg_label_to_id.keys()),
        placeholder="All regions...",
    )

    st.subheader("Job Type")
    type_label_to_id = {
        f"{v['en']}  ({v['he']})": k
        for k, v in sorted(JOB_TYPES.items(), key=lambda x: x[0])
    }
    selected_type_labels = st.multiselect(
        "Select job types",
        options=list(type_label_to_id.keys()),
        placeholder="All types...",
    )

    st.subheader("Keyword Search")
    search_text = st.text_input("Free text", placeholder="e.g. growth manager")

    st.divider()

    days_back = st.slider("Max posting age (days)", 1, 60, 14, help="Filter out jobs older than N days")
    max_pages = st.slider("Pages per search", 1, 20, 5, help="~20-30 jobs per page")

    st.divider()

    st.subheader("Translation")
    do_translate = st.checkbox("Translate Hebrew → English", value=True)
    if do_translate:
        translator_ok = get_translator()
        if translator_ok:
            st.success("Translation ready ✓")
        else:
            st.error("Translation unavailable — check internet connection")
            do_translate = False

    scrape_btn = st.button("🚀 Scrape AllJobs", type="primary", use_container_width=True)

    # ── WhatsApp settings ──
    st.divider()
    with st.expander("📱 WhatsApp Settings"):
        st.caption("Connects via WhatsApp Web — scan a QR code once, then sessions are saved.")
        wa_limit = st.slider("Messages to fetch", 50, 500, 200, step=50, key="wa_limit")
        wa_do_translate = st.checkbox("Translate to English after fetch", value=True, key="wa_translate")

    # ── Telegram settings ──
    st.divider()
    with st.expander("💬 Telegram Settings"):
        st.caption("Get API credentials at [my.telegram.org](https://my.telegram.org) → App configuration")
        tg_api_id = st.text_input(
            "API ID",
            value=os.getenv("TELEGRAM_API_ID", ""),
            key="tg_api_id_input",
        )
        tg_api_hash = st.text_input(
            "API Hash",
            type="password",
            value=os.getenv("TELEGRAM_API_HASH", ""),
            key="tg_api_hash_input",
        )
        tg_phone = st.text_input(
            "Phone number",
            placeholder="+972501234567",
            value=os.getenv("TELEGRAM_PHONE", ""),
            key="tg_phone_input",
        )
        tg_limit = st.slider("Messages to fetch", 50, 500, 200, step=50, key="tg_limit")
        tg_do_translate = st.checkbox("Translate to English after fetch", value=True, key="tg_translate")


# ─────────────────────────────────────────────
# Main area — tabs
# ─────────────────────────────────────────────

tab1, tab2, tab3 = st.tabs(["🏢 AllJobs.co.il", "💬 Telegram Jobs", "📱 WhatsApp Jobs"])


# ═══════════════════════════════════════════════
# TAB 1 — AllJobs.co.il
# ═══════════════════════════════════════════════

with tab1:

    # ── Scraping ──
    if scrape_btn:
        cat_ids = [cat_label_to_id[l] for l in selected_cat_labels] if selected_cat_labels else [None]
        reg_ids = [reg_label_to_id[l] for l in selected_reg_labels] if selected_reg_labels else [None]
        type_ids = [type_label_to_id[l] for l in selected_type_labels] if selected_type_labels else [None]

        if not selected_cat_labels and not selected_reg_labels and not selected_type_labels and not search_text:
            st.warning("⚠️ Please select at least one filter or enter a keyword.")
            st.stop()

        combos = []
        for cat in cat_ids:
            for reg in reg_ids:
                for typ in type_ids:
                    combos.append((cat, reg, typ))
                if not selected_type_labels:
                    break
        combos = combos[:15]

        scraper = AllJobsScraper(delay=1.5, verbose=False)
        seen_ids: set = set()
        all_jobs: list[dict] = []

        progress_bar = st.progress(0.0)
        status = st.empty()

        for i, (cat_id, reg_id, type_id) in enumerate(combos):
            cat_name = CATEGORIES.get(cat_id, {}).get("en", "All") if cat_id else "All"
            reg_name = REGIONS.get(reg_id, {}).get("en", "All") if reg_id else "All"
            status.info(f"Scraping: **{cat_name}** | Region: **{reg_name}** — pages 1-{max_pages}…")

            try:
                jobs = scraper.scrape_jobs(
                    category=cat_id,
                    region=reg_id,
                    job_type=type_id,
                    search_text=search_text or None,
                    max_pages=max_pages,
                )
            except Exception as e:
                st.warning(f"Scrape error ({cat_name}): {e}")
                jobs = []

            for job in jobs:
                jid = job.get("job_id")
                if jid and jid not in seen_ids:
                    age = parse_days_ago(job.get("date_posted", ""))
                    if age <= days_back:
                        seen_ids.add(jid)
                        all_jobs.append(job)

            progress_bar.progress((i + 1) / len(combos))

        status.empty()
        progress_bar.empty()

        if not all_jobs:
            st.info("No jobs found. Try expanding the date range, adding more categories, or using different keywords.")
            st.stop()

        if do_translate:
            trans_bar = st.progress(0.0)
            trans_status = st.empty()
            fields_to_translate = ["title", "company", "location", "job_type", "description", "requirements"]
            for i, job in enumerate(all_jobs):
                trans_status.info(f"Translating job {i + 1} / {len(all_jobs)}…")
                for field in fields_to_translate:
                    val = job.get(field, "")
                    job[f"{field}_en"] = translate_text(val) if val else ""
                trans_bar.progress((i + 1) / len(all_jobs))
                if (i + 1) % 10 == 0:
                    time.sleep(0.5)
            trans_bar.empty()
            trans_status.empty()

        st.session_state["jobs"] = all_jobs
        st.session_state["translated"] = do_translate
        st.success(f"✅ Done! Found **{len(all_jobs)}** jobs.")

    # ── Results ──
    if "jobs" in st.session_state and st.session_state["jobs"]:
        jobs = st.session_state["jobs"]
        translated = st.session_state.get("translated", False)

        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Total Jobs", len(jobs))
        col2.metric("Unique Companies", len({j.get("company", "") for j in jobs if j.get("company")}))
        col3.metric("Posted Today / Yesterday", sum(1 for j in jobs if parse_days_ago(j.get("date_posted", "")) <= 1))
        col4.metric("Within 7 days", sum(1 for j in jobs if parse_days_ago(j.get("date_posted", "")) <= 7))

        st.divider()

        excel_bytes = jobs_to_excel(jobs, translated, source_col="direct_url")
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(
            label="📥 Download Excel (.xlsx)",
            data=excel_bytes,
            file_name=f"alljobs_{ts}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )

        st.divider()

        filter_col1, filter_col2 = st.columns([3, 1])
        with filter_col1:
            search_filter = st.text_input("🔎 Filter displayed results", placeholder="Search titles, companies…", key="aj_filter")
        with filter_col2:
            sort_by = st.selectbox("Sort by", ["Newest first", "Oldest first", "Company A-Z"], key="aj_sort")

        display_jobs = jobs
        if search_filter:
            q = search_filter.lower()
            display_jobs = [
                j for j in jobs
                if q in (j.get("title_en") or j.get("title") or "").lower()
                or q in (j.get("company_en") or j.get("company") or "").lower()
                or q in (j.get("description_en") or j.get("description") or "").lower()
            ]

        if sort_by == "Newest first":
            display_jobs = sorted(display_jobs, key=lambda j: parse_days_ago(j.get("date_posted", "")))
        elif sort_by == "Oldest first":
            display_jobs = sorted(display_jobs, key=lambda j: parse_days_ago(j.get("date_posted", "")), reverse=True)
        elif sort_by == "Company A-Z":
            display_jobs = sorted(display_jobs, key=lambda j: (j.get("company_en") or j.get("company") or "").lower())

        st.subheader(f"Showing {len(display_jobs)} jobs")
        render_job_cards(display_jobs, translated, url_field="direct_url")


# ═══════════════════════════════════════════════
# TAB 2 — Telegram Jobs
# ═══════════════════════════════════════════════

with tab2:

    st.subheader("Hebrew Job Posts from Telegram")
    st.caption("Reads from a Telegram group you're a member of, structures each post with Claude AI, then translates to English.")

    # ── Authentication flow ──
    creds_ok = tg_api_id and tg_api_hash and tg_phone

    if not creds_ok:
        st.info("Fill in your Telegram credentials in the sidebar **Telegram Settings** expander to get started.")
        st.stop()

    # Fresh fetcher each rerun — avoids asyncio event loop conflicts with Streamlit
    fetcher = TelegramJobFetcher(int(tg_api_id), tg_api_hash, tg_phone)

    for key, default in [
        ("tg_authenticated", False),
        ("tg_phone_code_hash", None),
        ("tg_dialogs", []),
        ("tg_jobs", []),
        ("tg_translated", False),
    ]:
        if key not in st.session_state:
            st.session_state[key] = default

    # ── Step 1: Connect ──
    if not st.session_state["tg_authenticated"]:
        st.markdown("**Step 1 — Connect your Telegram account**")

        connect_btn = st.button("🔗 Connect & Send Code", type="primary")
        if connect_btn:
            try:
                if fetcher.is_authorized():
                    st.session_state["tg_authenticated"] = True
                    st.session_state["tg_dialogs"] = fetcher.list_dialogs()
                    st.success("Already authenticated via saved session!")
                    st.rerun()
                else:
                    phone_code_hash = fetcher.send_code()
                    st.session_state["tg_phone_code_hash"] = phone_code_hash
                    st.info("A verification code was sent to your Telegram app. Enter it below.")
            except Exception as e:
                st.error(f"Connection error: {e}")

        # ── Step 2: Enter code + optional 2FA ──
        if st.session_state["tg_phone_code_hash"]:
            st.markdown("**Step 2 — Enter the verification code**")
            code_col, btn_col = st.columns([3, 1])
            with code_col:
                tg_code = st.text_input("Verification code", max_chars=10, key="tg_code_input")
            with btn_col:
                st.write("")
                verify_btn = st.button("✅ Verify", type="primary")

            tg_2fa = st.text_input("2FA password (if enabled)", type="password", key="tg_2fa_input")

            if verify_btn and tg_code:
                try:
                    fetcher.sign_in(tg_code, st.session_state["tg_phone_code_hash"], password=tg_2fa)
                    st.session_state["tg_authenticated"] = True
                    st.session_state["tg_dialogs"] = fetcher.list_dialogs()
                    st.success("Authenticated! Session saved for future use.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Verification failed: {e}")

    # ── Step 3: Fetch posts ──
    else:
        dialogs = st.session_state["tg_dialogs"]
        if not dialogs:
            try:
                st.session_state["tg_dialogs"] = fetcher.list_dialogs()
                dialogs = st.session_state["tg_dialogs"]
            except Exception as e:
                st.error(f"Could not load groups: {e}")
                st.session_state["tg_authenticated"] = False
                st.rerun()

        if not dialogs:
            st.warning("No groups or channels found for this account.")
        else:
            dialog_names = [d["name"] for d in dialogs]
            selected_name = st.selectbox("Select Telegram group / channel", dialog_names, key="tg_dialog_select")
            selected_dialog = next((d for d in dialogs if d["name"] == selected_name), None)

            fetch_col, reset_col = st.columns([3, 1])
            with fetch_col:
                fetch_btn = st.button("📥 Fetch Posts", type="primary", use_container_width=True)
            with reset_col:
                if st.button("🔓 Disconnect", use_container_width=True):
                    st.session_state["tg_authenticated"] = False
                    st.session_state["tg_dialogs"] = []
                    st.session_state["tg_jobs"] = []
                    st.rerun()

            if fetch_btn and selected_dialog:
                try:
                    tg_progress = st.progress(0.0)
                    tg_status = st.empty()
                    tg_status.info(f"Fetching up to {tg_limit} messages from **{selected_name}**…")

                    jobs = fetcher.fetch_messages(selected_dialog["id"], limit=tg_limit)
                    tg_progress.progress(0.5)

                    if not jobs:
                        st.warning("No text messages found in this group.")
                    else:
                        if tg_do_translate:
                            tg_status.info(f"Translating {len(jobs)} posts to English…")
                            fields_to_translate = ["title", "description"]
                            for i, job in enumerate(jobs):
                                for field in fields_to_translate:
                                    val = job.get(field, "")
                                    job[f"{field}_en"] = translate_text(val) if val else ""
                                tg_progress.progress(0.5 + (i + 1) / len(jobs) * 0.5)
                                if (i + 1) % 10 == 0:
                                    time.sleep(0.5)

                        st.session_state["tg_jobs"] = jobs
                        st.session_state["tg_translated"] = tg_do_translate

                    tg_progress.empty()
                    tg_status.empty()
                    st.success(f"✅ Fetched **{len(jobs)}** posts from {selected_name}.")

                except Exception as e:
                    st.error(f"Fetch error: {e}")

        # ── Results ──
        tg_jobs = st.session_state.get("tg_jobs", [])
        tg_translated = st.session_state.get("tg_translated", False)

        if tg_jobs:
            st.divider()

            tc1, tc2, tc3 = st.columns(3)
            tc1.metric("Total Posts", len(tg_jobs))
            tc2.metric("Unique Senders", len({j.get("sender", "") for j in tg_jobs if j.get("sender")}))
            if tg_jobs:
                dates = [j.get("date_ts", 0) for j in tg_jobs if j.get("date_ts")]
                if dates:
                    oldest = datetime.fromtimestamp(min(dates)).strftime("%Y-%m-%d")
                    tc3.metric("Oldest post", oldest)

            excel_bytes = jobs_to_excel(tg_jobs, tg_translated, source_col="source_url")
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            st.download_button(
                label="📥 Download Excel (.xlsx)",
                data=excel_bytes,
                file_name=f"telegram_jobs_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )

            st.divider()

            tg_filter_col, tg_sort_col = st.columns([3, 1])
            with tg_filter_col:
                tg_search = st.text_input("🔎 Filter posts", placeholder="Search titles, companies…", key="tg_filter")
            with tg_sort_col:
                tg_sort = st.selectbox("Sort by", ["Newest first", "Oldest first", "Company A-Z"], index=0, key="tg_sort")

            # Always default to newest first
            display_tg = sorted(tg_jobs, key=lambda j: j.get("date_ts", 0), reverse=True)

            if tg_search:
                q = tg_search.lower()
                display_tg = [
                    j for j in display_tg
                    if q in (j.get("title_en") or j.get("title") or "").lower()
                    or q in (j.get("company_en") or j.get("company") or "").lower()
                    or q in (j.get("description_en") or j.get("description") or "").lower()
                ]

            if tg_sort == "Oldest first":
                display_tg = sorted(display_tg, key=lambda j: j.get("date_ts", 0))
            elif tg_sort == "Company A-Z":
                display_tg = sorted(display_tg, key=lambda j: (j.get("company_en") or j.get("company") or "").lower())

            st.subheader(f"Showing {len(display_tg)} posts")
            render_job_cards(display_tg, tg_translated, url_field="source_url", max_desc=10000)


# ═══════════════════════════════════════════════
# TAB 3 — WhatsApp Jobs
# ═══════════════════════════════════════════════

with tab3:

    st.subheader("Hebrew Job Posts from WhatsApp")
    st.caption("Reads messages from a WhatsApp group you're a member of. Scan a QR code once — session is saved for future runs.")

    # Init session state
    for key, default in [
        ("wa_server_ok", False),
        ("wa_status", "disconnected"),
        ("wa_groups", []),
        ("wa_jobs", []),
        ("wa_translated", False),
    ]:
        if key not in st.session_state:
            st.session_state[key] = default

    # ── Ensure bridge server is running ──
    if not st.session_state["wa_server_ok"]:
        with st.spinner("Starting WhatsApp bridge server…"):
            ok = wa.ensure_server_running()
        if ok:
            st.session_state["wa_server_ok"] = True
        else:
            st.error("Could not start the WhatsApp bridge server. Make sure Node.js is installed and `node_modules` exists in the app directory.")
            st.stop()

    # Always refresh status on every rerun
    srv = wa.get_status()
    st.session_state["wa_status"] = srv.get("status", "disconnected")
    wa_status = st.session_state["wa_status"]

    # ── Step 1: Not connected ──
    if wa_status in ("disconnected", "error"):
        st.markdown("**Step 1 — Connect your WhatsApp account**")
        st.info("Click the button below — a QR code will appear. Scan it with WhatsApp on your phone (Settings → Linked Devices → Link a Device).")

        if st.button("📱 Connect WhatsApp", type="primary"):
            wa.connect()
            st.rerun()

    # ── QR pending: show QR and auto-refresh ──
    elif wa_status in ("initializing", "qr_pending"):
        qr_b64 = srv.get("qr")
        if qr_b64:
            st.markdown("**Scan this QR code with WhatsApp on your phone**")
            st.image(qr_b64, width=280)
            st.caption("WhatsApp → Settings → Linked Devices → Link a Device")
        else:
            st.info("Initialising WhatsApp client… please wait.")

        col_refresh, col_cancel = st.columns([2, 1])
        with col_refresh:
            if st.button("🔄 Refresh Status", use_container_width=True):
                st.rerun()
        with col_cancel:
            if st.button("✖ Cancel", use_container_width=True):
                wa.disconnect()
                st.session_state["wa_status"] = "disconnected"
                st.rerun()

        # Auto-refresh every 3 s while waiting for scan
        time.sleep(3)
        st.rerun()

    # ── Connected: group picker + fetch ──
    elif wa_status == "connected":
        st.success("✅ WhatsApp connected")

        # Load groups once per session
        if not st.session_state["wa_groups"]:
            try:
                with st.spinner("Loading groups…"):
                    st.session_state["wa_groups"] = wa.list_groups()
            except RuntimeError as e:
                if "DETACHED_FRAME" in str(e):
                    st.warning("WhatsApp session lost (browser was restarted). Click **Reconnect** to re-link.")
                    if st.button("🔄 Reconnect WhatsApp", type="primary"):
                        wa.connect()
                        st.session_state["wa_status"] = "initializing"
                        st.rerun()
                    st.stop()
                else:
                    st.error(str(e))

        groups = st.session_state["wa_groups"]

        if not groups:
            st.warning("No WhatsApp groups found for this account.")
        else:
            group_names = [f"{g['name']}  ({g['participants']} members)" for g in groups]
            selected_idx = st.selectbox(
                "Select WhatsApp group",
                range(len(group_names)),
                format_func=lambda i: group_names[i],
                key="wa_group_select",
            )
            selected_group = groups[selected_idx]

            fetch_col, disc_col = st.columns([3, 1])
            with fetch_col:
                fetch_btn = st.button("📥 Fetch Messages", type="primary", use_container_width=True)
            with disc_col:
                if st.button("🔓 Disconnect", use_container_width=True):
                    wa.disconnect()
                    st.session_state.update({
                        "wa_status": "disconnected",
                        "wa_groups": [],
                        "wa_jobs": [],
                    })
                    st.rerun()

            if fetch_btn:
                try:
                    wa_progress = st.progress(0.0)
                    wa_status_msg = st.empty()
                    wa_status_msg.info(f"Fetching up to {wa_limit} messages from **{selected_group['name']}**…")

                    jobs = wa.fetch_messages(selected_group["id"], limit=wa_limit)
                    wa_progress.progress(0.5)

                    if not jobs:
                        st.warning("No messages found in this group (minimum 20 characters per message).")
                    else:
                        if wa_do_translate:
                            wa_status_msg.info(f"Translating {len(jobs)} messages to English…")
                            for i, job in enumerate(jobs):
                                for field in ["title", "description"]:
                                    val = job.get(field, "")
                                    job[f"{field}_en"] = translate_text(val) if val else ""
                                wa_progress.progress(0.5 + (i + 1) / len(jobs) * 0.5)
                                if (i + 1) % 10 == 0:
                                    time.sleep(0.5)

                        st.session_state["wa_jobs"] = jobs
                        st.session_state["wa_translated"] = wa_do_translate

                    wa_progress.empty()
                    wa_status_msg.empty()
                    st.success(f"✅ Fetched **{len(jobs)}** messages from {selected_group['name']}.")

                except Exception as e:
                    st.error(f"Fetch error: {e}")

        # ── Results ──
        wa_jobs = st.session_state.get("wa_jobs", [])
        wa_translated = st.session_state.get("wa_translated", False)

        if wa_jobs:
            st.divider()

            wc1, wc2, wc3 = st.columns(3)
            wc1.metric("Total Messages", len(wa_jobs))
            wc2.metric("Unique Senders", len({j.get("sender", "") for j in wa_jobs if j.get("sender")}))
            dates = [j.get("date_ts", 0) for j in wa_jobs if j.get("date_ts")]
            if dates:
                wc3.metric("Oldest message", datetime.fromtimestamp(min(dates)).strftime("%Y-%m-%d"))

            excel_bytes = jobs_to_excel(wa_jobs, wa_translated, source_col="source_url")
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            st.download_button(
                label="📥 Download Excel (.xlsx)",
                data=excel_bytes,
                file_name=f"whatsapp_jobs_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )

            st.divider()

            wa_fcol, wa_scol = st.columns([3, 1])
            with wa_fcol:
                wa_search = st.text_input("🔎 Filter messages", placeholder="Search text, senders…", key="wa_filter")
            with wa_scol:
                wa_sort = st.selectbox("Sort by", ["Newest first", "Oldest first", "Sender A-Z"], index=0, key="wa_sort")

            display_wa = sorted(wa_jobs, key=lambda j: j.get("date_ts", 0), reverse=True)

            if wa_search:
                q = wa_search.lower()
                display_wa = [
                    j for j in display_wa
                    if q in (j.get("title_en") or j.get("title") or "").lower()
                    or q in (j.get("sender") or "").lower()
                    or q in (j.get("description_en") or j.get("description") or "").lower()
                ]

            if wa_sort == "Oldest first":
                display_wa = sorted(display_wa, key=lambda j: j.get("date_ts", 0))
            elif wa_sort == "Sender A-Z":
                display_wa = sorted(display_wa, key=lambda j: (j.get("sender") or "").lower())

            st.subheader(f"Showing {len(display_wa)} messages")
            render_job_cards(display_wa, wa_translated, url_field="source_url", max_desc=10000)
